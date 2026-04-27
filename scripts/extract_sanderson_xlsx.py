"""
Extract Sanderson market-adjusted rankings and target exposures from
ALL 2026 ROOKIE RANKINGS.xlsx, SF TEP tab.

Market-adj ranks (cols G/H) reset per round (1-12 per round).
We unwind to absolute rank by adding (round_num - 1) * 12 per-round.

Target exposures (cols J/K) are sorted by ADP — separate player order.
We match them to market-adj players by abbreviated name.

Output: scripts/output/sanderson_xlsx.json
Keys are abbreviated names (e.g. "J. Love") for fuzzy matching downstream.
"""
import json, os
import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'src', 'ALL 2026 ROOKIE RANKINGS.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'output', 'sanderson_xlsx.json')

ROUND_SIZE = 12


def extract():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['SF TEP']

    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

    # ── Market-adj ranks (cols G=idx6, H=idx7) ───────────────────────────────
    mkt_ranks = {}  # abbr_name -> {sanderson_mkt_rank, sanderson_mkt_round}
    round_num = 0
    for row in rows:
        rank_val = row[6]   # col G
        name     = row[7]   # col H
        if rank_val is None or name is None:
            continue
        rank_int = int(rank_val)
        if rank_int == 1:
            round_num += 1
        abs_rank = rank_int + (round_num - 1) * ROUND_SIZE
        mkt_ranks[name.strip()] = {
            'sanderson_mkt_rank': abs_rank,
            'sanderson_mkt_round': round_num,
        }

    # ── Target exposures (cols J=idx9, K=idx10) ──────────────────────────────
    exposures = {}  # abbr_name -> exposure string
    for row in rows:
        exp_val  = row[9]   # col J
        exp_name = row[10]  # col K
        if exp_val is None or exp_name is None:
            continue
        exposures[exp_name.strip()] = str(exp_val).strip()

    # ── Merge: attach exposure to each market-adj player ─────────────────────
    def _last(name):
        """Return last-name portion (after the first-initial dot, if present)."""
        parts = name.strip().split('.')
        return parts[-1].strip().lower() if len(parts) > 1 else name.strip().lower()

    result = {}
    for name, rank_data in mkt_ranks.items():
        entry = dict(rank_data)
        # Exact match first; fall back to last-name match
        exp = exposures.get(name)
        if exp is None:
            ln = _last(name)
            for ename, eval_ in exposures.items():
                if _last(ename) == ln:
                    exp = eval_
                    break
        entry['sanderson_exposure'] = exp
        result[name] = entry

    # Also add exposure-only players (not in market-adj list) so build_data
    # can pick up exposures for players not ranked by Sanderson this week.
    for name, exp in exposures.items():
        if name not in result:
            result[name] = {'sanderson_mkt_rank': None, 'sanderson_mkt_round': None,
                            'sanderson_exposure': exp}

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(result)} entries to {OUTPUT_PATH}")
    print("\nMarket-adj ranked players (first 15):")
    ranked = [(v['sanderson_mkt_rank'], k, v['sanderson_exposure'])
              for k, v in result.items() if v['sanderson_mkt_rank']]
    for rank, name, exp in sorted(ranked)[:15]:
        print(f"  {rank:2d}. {name:<25} exp={exp}")
    return result


if __name__ == '__main__':
    extract()
